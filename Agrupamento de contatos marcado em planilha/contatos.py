import openpyxl as xl
import datetime
now = datetime.datetime.now()

planilha = xl.load_workbook(filename = '40.000Contatos.xlsx') 
worksheets_fonte = planilha.worksheets[0] 
mc = worksheets_fonte.max_row

pl2 = xl.load_workbook(filename = 'manter.xlsx') 
worksheets_importar = pl2.active
#worksheets_importar.title = "Import" 

#pl3 = xl.load_workbook(filename = 'deletar.xlsx') 
worksheets_excluir = pl3.active 

def get_colunas(worksheets):
	grupo = "Group:"
	col = []
	for row in worksheets.rows:
		for cell in row:
			if(cell.value == None):
				print("fim")
				break;
			elif (grupo in cell.value):
				print(cell.value)
				col.append(cell.coordinate[:2])		
		break
	print(col)
	print(len(col))
	return col

def checar_grupo(worksheets, lista):
	pertence_a_grupo = []
	linha = 0
	excluir = 0
	manter = 0

	for i in range (1, mc+1): 
		pertence_a_grupo.clear()
		linha+=1
		for j in lista:
			index = (j + str(i))
			c = worksheets[index]
			pertence_a_grupo.append(True) if c.value == "True" else pertence_a_grupo.append(False)

		if any(pertence_a_grupo):
			escrever_planilha(worksheets, worksheets_importar, linha)
			manter +=1
			
		else:
			escrever_planilha(worksheets, worksheets_excluir, linha)
			excluir +=1
		
	print("Excluidos: ", excluir, "Mantidos: ", manter)

def salvar_cabecalho(worksheets1, worksheets2):	
	mc = worksheets1.max_column 
	for j in  range (1, mc + 1): 
		c = worksheets1.cell(row = 1, column = j) 
		worksheets2.cell(row = 1, column = j).value = c.value

def escrever_planilha(worksheets1, worksheets2, linha):	
	mc = worksheets1.max_column  
	list = []
	for j in  range (1, mc + 1): 
		c = worksheets1.cell(row = linha, column = j) 
		list.append(c.value)
	worksheets2.append(list)	

def cronometro():
	now2 = datetime.datetime.now()

	tempo = now2 - now
	print("Tempo até aqui:", tempo)

def main():
	print("Máximo de linhas: ",mc)
	col2 = get_colunas(worksheets_fonte)
	print(col2)	
	cronometro()
	blz = input("Vamos seguir, blz?")

	salvar_cabecalho(worksheets_fonte, worksheets_importar)
	#salvar_cabecalho(worksheets_fonte, worksheets_excluir)
	checar_grupo(worksheets_fonte, col2)
	pl2.save('manter.xlsx')
	pl3.save('deletar.xlsx')
	
	planilha.close()
	pl2.close()
	pl3.close()



if __name__ == '__main__':
	main() 

	cronometro()

